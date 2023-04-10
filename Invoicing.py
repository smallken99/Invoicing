import sys
from PyQt5.QtWidgets import QApplication, QMessageBox, QHBoxLayout, QLabel, QWidget, QPushButton, QLineEdit, QVBoxLayout, QComboBox, QTabWidget
import openpyxl
from PyQt5 import QtGui
from PyQt5.QtCore import Qt

class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('My Window')


        options = []

        # 從檔案讀取選項
        with open('products.txt', 'r') as f:
            for line in f:
                # 每行一個選項，存儲在list中
                options.append(line.strip())

        self.combos = []
        for i in range(5) :
            combo = QComboBox(self)
            for option in options :
                combo.addItem(option)
                combo.setFont(QtGui.QFont('Arial', 12))
            self.combos.append(combo)

        # 創建五個下拉式選單的陣列
        # 創建五個數字的輸入欄位
        self.amount_edit = []
        for i in range(5):
            amount = QLineEdit(self)
            amount.setValidator(QtGui.QIntValidator())
            self.amount_edit.append(amount)


        # 創建五個數量下拉框控件陣列
        self.quantity = []
        for i in range(5):
            quantity = QComboBox(self)
            for j in range(0, 11):
                quantity.addItem(str(j))
                quantity.setFont(QtGui.QFont('Arial', 12))
            self.quantity.append(quantity)



        # 创建水平布局，并将各控件添加到水平布局中
        self.hbox = []
        font = QtGui.QFont("Arial", 12)  # 設定字型為Arial，字體大小為14
        for i in range(5):
            hbox = QHBoxLayout()
            label0 = QLabel(f'{i+1} 進貨品項', self)          
            label0.setFont(font)
            hbox.addWidget(label0)
            hbox.addWidget(self.combos[i])
            label = QLabel("金額", self)          
            label.setFont(font)
            hbox.addWidget(label)
            self.amount_edit[i].setFixedWidth(100)
            self.amount_edit[i].setFont(QtGui.QFont('Arial', 12))
            hbox.addWidget(self.amount_edit[i])
            label2 = QLabel("數量", self)
            label2.setFont(font)
            hbox.addWidget(label2)
            hbox.addWidget(self.quantity[i])
            self.hbox.append(hbox)

        # 创建進項稅額
        hbox_tax = QHBoxLayout()
        hbox.setAlignment(Qt.AlignLeft)
        labelTax =  QLabel("進項稅額", self)  
        labelTax.setFont(font)
        hbox_tax.addWidget(labelTax)
        self.taxAmount = QLineEdit(self)
        self.taxAmount.setFixedWidth(100)
        self.taxAmount.setFont(QtGui.QFont('Arial', 12))
        self.taxAmount.setValidator(QtGui.QIntValidator())
        hbox_tax.addWidget(self.taxAmount)
        hbox_tax.addWidget(QLabel("-------------------------------------------------------------------------------------"))

        # 送出按鈕
        self.button = QPushButton('送出')
        self.button.setFont(QtGui.QFont('Arial', 12))

        # 清除按鈕
        #self.button2 = QPushButton('清除')
        #self.button2.setFont(QtGui.QFont('Arial', 12))


        # 创建垂直布局，并将水平布局添加到垂直布局中
        self.vbox1 = QVBoxLayout()
        for i in range(5):
            self.vbox1.addLayout(self.hbox[i])

        #
        self.vbox1.addLayout(hbox_tax)
        self.vbox1.addWidget(self.button)
        #self.vbox1.addWidget(self.button2)

        tab1 = QWidget()
        tab1.setLayout(self.vbox1)


        # 創建第二個頁籤，和第一個頁籤相同
        tab2 = QWidget()
        vbox2 = QVBoxLayout()

        tab2.setLayout(vbox2)

        # 創建標籤式選項卡，添加兩個頁籤
        tabs = QTabWidget()
        tabs.addTab(tab1, '進貨')
        tabs.addTab(tab2, '出貨')

        # 設置窗口布局
        vbox = QVBoxLayout()
        vbox.addWidget(tabs)
        self.setLayout(vbox)

        # 設置按鈕點擊事件處理函數
        self.button.clicked.connect(self.buttonClicked)
        #self.button2.clicked.connect(self.resetData)


    # 找出指定工作表中 A 欄最後一個非空單元格的位置，並回傳下一個單元格的索引    
    def find_last_empty_row(self,sh):
        max_row = sh.max_row
        for i in range(max_row, 0, -1):
            if sh.cell(row=i, column=1).value is not None:
                max_row = i
                break        
        return max_row


    def buttonClicked(self):
        # 按鈕點擊事件處理函數 讀取介面內容
        product_text = []
        amount_text = []
        quantity_text = []
        TotalAmount = 0
        for i in range(5):
            product_text.append(self.combos[i].currentText())
            amount_text.append(self.amount_edit[i].text())
            quantity_text.append(self.quantity[i].currentText())

        msg = ""
        for i in range(5) :
            if amount_text[i].strip() and int(quantity_text[i]) > 0 :
                msg += f'{product_text[i]} ,金額 {amount_text[i]}, 數量 {quantity_text[i]} \n'
                TotalAmount += int(amount_text[i])
        if self.taxAmount.text().strip() :
            TotalAmount += int(self.taxAmount.text())  
        msg += f'進項稅額　{self.taxAmount.text()} \n'
        msg += f'現　　金　{TotalAmount} \n'
        msg += "請問是否確認存檔?"
        msg_box = QMessageBox()
        msg_box.setWindowTitle('確認')
        msg_box.setText(msg)
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        choice = msg_box.exec_()

        # If user clicks "Yes", save Excel file
        if choice == QMessageBox.No:
             return
        else:
            pass
    
        # 打開 Excel 文件
        try:
            wb = openpyxl.load_workbook('112財報資料.xlsx', read_only=False)
        except FileNotFoundError:
            print('Excel 文件不存在')
            return

        # 選擇工作表
        sheet = wb['112日記帳']

        # 找到 A 欄中最後一個有值的單元格
        max_row = self.find_last_empty_row(sheet)

        # 在最後一行的下一行插入新資料
        for i in range(5):
            if amount_text[i].strip() and int(quantity_text[i]) > 0:
                max_row += 1
                sheet.cell(row=max_row, column=1).value = "-"
                sheet.cell(row=max_row, column=2).value = product_text[i]
                sheet.cell(row=max_row, column=4).value = int(amount_text[i])
   
        
        # 進項稅額
        tax_amount = self.taxAmount.text()
        if tax_amount.strip() and int(tax_amount) > 0 :
            max_row += 1
            sheet.cell(row=max_row, column=1).value = "-"
            sheet.cell(row=max_row, column=2).value = "進項稅額" 
            sheet.cell(row=max_row, column=4).value = int(tax_amount)

        # 現金欄位
        max_row += 1
        sheet.cell(row=max_row, column=1).value = "-"
        sheet.cell(row=max_row, column=3).value = "現金" 
        sheet.cell(row=max_row, column=5).value = TotalAmount      


        # 選擇工作表 存貨
        sheet2 = wb['存貨']
        # 找到 A 欄中最後一個有值的單元格
        max_row = self.find_last_empty_row(sheet2)
        # 在最後一行的下一行插入新資料
        for i in range(5):
            if int(quantity_text[i]) > 0:
                max_row += 1
                sheet2.cell(row=max_row, column=1).value = "-"
                sheet2.cell(row=max_row, column=2).value = product_text[i]
                sheet2.cell(row=max_row, column=4).value = int(amount_text[i])



        # 選擇工作表 進項稅額
        sheet3 = wb['進項稅額']
        # 找到 A 欄中最後一個有值的單元格
        max_row = self.find_last_empty_row(sheet3)
        if tax_amount.strip() and int(tax_amount) > 0 :
            max_row += 1
            sheet3.cell(row=max_row, column=1).value = "-"
            sheet3.cell(row=max_row, column=2).value = "進項稅額" 
            sheet3.cell(row=max_row, column=4).value = int(self.taxAmount.text())     


        # 選擇工作表 現金
        sheet4 = wb['現金']
        # 找到 A 欄中最後一個有值的單元格
        max_row = self.find_last_empty_row(sheet4)

        max_row += 1
        sheet4.cell(row=max_row, column=1).value = "-"
        sheet4.cell(row=max_row, column=3).value = "現金" 
        sheet4.cell(row=max_row, column=5).value = TotalAmount  


        # 保存文件
        wb.save('112財報資料.xlsx')
        print('已將資料寫入 Excel 文件')
        QMessageBox.information(self, 'Save Excel', '資料寫入成功')
        self.resetData() # 重置資料

    # 重置資料
    def resetData(self):
        for i in range(5):
            self.combos[i].setCurrentIndex(0)
            self.amount_edit[i].clear()
            self.quantity[i].setCurrentIndex(0)
            self.taxAmount.clear()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())