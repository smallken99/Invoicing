from PyQt5.QtWidgets import QApplication, QMessageBox, QHBoxLayout, QLabel, QWidget, QPushButton, QLineEdit, QVBoxLayout, QComboBox, QTabWidget
from PyQt5 import QtGui
from PyQt5.QtCore import Qt
import openpyxl
import my_module

class Tab2Widget(QWidget):
    def __init__(self, options, parent):
        super().__init__(parent)
        self.options = options
        #print(options)
        self.initUI()


    def initUI(self):
        self.saleCostAmount = 0
        self.saleCashAmount = 0
        # 創建n個產品項目清單,下拉式選單
        self.optionItems = []
        for i in range(5) :
            combosItem = QComboBox(self)
            for option in self.options :
                combosItem.addItem(option)
                combosItem.setFont(QtGui.QFont('Arial', 12))
            self.optionItems.append(combosItem)

        # 創建n個數字的輸入欄位
        self.amount_edit = []
        for i in range(5):
            amount = QLineEdit(self)
            amount.setValidator(QtGui.QIntValidator())
            amount.textChanged.connect(self.updateCostSum)
            self.amount_edit.append(amount)

        # 創建n個數量下式選單
        self.quantity = []
        for i in range(5):
            quantity = QComboBox(self)
            for j in range(0, 11):
                quantity.addItem(str(j))
                quantity.setFont(QtGui.QFont('Arial', 12))
            quantity.currentIndexChanged.connect(self.updateCostSum)
            self.quantity.append(quantity)

  

        font = QtGui.QFont("Arial", 12)  # 設定字型為Arial，字體大小為12


        # 創建發票號碼 label
        hbox_invoice = QHBoxLayout()
        hbox_invoice.setAlignment(Qt.AlignLeft)
        self.labelinvoice =  QLabel("發票號碼", self)  
        self.labelinvoice.setFont(font)
        # 創建發票號碼輸入欄位
        self.invoice = QLineEdit(self)
        self.invoice.setFixedWidth(130)
        self.invoice.setMaxLength(10)
        self.invoice.setFont(font)        
        hbox_invoice.addWidget(self.labelinvoice)
        hbox_invoice.addWidget(self.invoice)
        hbox_invoice.addWidget(QLabel("                                                                                                                   "))


        # 創建銷售成本 label
        hbox_cost = QHBoxLayout()
        hbox_cost.setAlignment(Qt.AlignLeft)
        self.labelCost =  QLabel("銷售成本", self)  
        self.labelCost.setFont(font)
        hbox_cost.addWidget(self.labelCost)
        hbox_cost.addWidget(QLabel("    -----------------------------------------------------------------",self))


        # 創建水平布局，並將各元件加到水平布局中
        self.hbox = []   

        for i in range(5):
            hbox = QHBoxLayout()
            # 品項 label
            label0 = QLabel(f'{i+1} 品項', self)          
            label0.setFont(font)
            hbox.addWidget(label0)
            # 品項元件,下拉式選單
            hbox.addWidget(self.optionItems[i])
            # 金額 label
            label = QLabel("金額", self)          
            label.setFont(font)
            hbox.addWidget(label)
            # 金額輸入欄位 元件
            self.amount_edit[i].setFixedWidth(100)
            self.amount_edit[i].setFont(font)
            hbox.addWidget(self.amount_edit[i])
            # 數量 label
            label2 = QLabel("數量", self)
            label2.setFont(font)
            hbox.addWidget(label2)
            # 數量下拉式選單
            hbox.addWidget(self.quantity[i])
            self.hbox.append(hbox) 

        # 創建現金 label
        hbox_cash = QHBoxLayout()
        hbox_cash.setAlignment(Qt.AlignLeft)
        self.labelCash =  QLabel("現金", self)  
        self.labelCash.setFont(font)
        hbox_cash.addWidget(self.labelCash)
        hbox_cash.addWidget(QLabel("    -----------------------------------------------------------------",self))

        # 創建銷收入
        hbox_income = QHBoxLayout()
        hbox_income.addWidget(QLabel("----------------------------------------------------------------------------------------------",self))
        # 銷貨收入 label
        labelIncome = QLabel("銷貨收入",self)
        labelIncome.setFont(font)
        # 銷貨收入輸入欄位
        self.incomeAmount = QLineEdit(self)
        self.incomeAmount.setFixedWidth(100)
        self.incomeAmount.setFont(font)
        self.incomeAmount.setValidator(QtGui.QIntValidator())
        self.incomeAmount.textChanged.connect(self.updateCashSum)
        hbox_income.addWidget(labelIncome)
        hbox_income.addWidget(self.incomeAmount)


        # 創建銷項稅額
        hbox_tax = QHBoxLayout()
        hbox_tax.addWidget(QLabel("----------------------------------------------------------------------------------------------",self))
        # 銷項稅額 label
        labelTax = QLabel("銷項稅額",self)
        labelTax.setFont(font)        
        # 銷項稅額輸入欄位
        self.taxAmount = QLineEdit(self)
        self.taxAmount.setFixedWidth(100)
        self.taxAmount.setFont(font)
        self.taxAmount.setValidator(QtGui.QIntValidator())
        self.taxAmount.textChanged.connect(self.updateCashSum)
        hbox_tax.addWidget(labelTax)
        hbox_tax.addWidget(self.taxAmount)

        # 創建送出按鈕
        self.button = QPushButton('送出')
        self.button.setFont(font)

        # 創建垂直布局
        self.vbox1 = QVBoxLayout()
        self.vbox1.addLayout(hbox_invoice)
        self.vbox1.addLayout(hbox_cost)
        for i in range(5):
            self.vbox1.addLayout(self.hbox[i])
        self.vbox1.addLayout(hbox_cash)
        self.vbox1.addLayout(hbox_income)
        self.vbox1.addLayout(hbox_tax)
        self.vbox1.addWidget(self.button)

        self.setLayout(self.vbox1)

        # 設置按鈕點擊事件處理函數
        self.button.clicked.connect(self.buttonClicked)

    def updateCostSum(self):
        self.saleCostAmount = 0
        for i in range(5) :
            if self.amount_edit[i].text().strip() and int(self.quantity[i].currentText()) > 0:
                self.saleCostAmount += int(self.amount_edit[i].text())
        # 在標籤中顯示總和
        self.labelCost.setText(f'銷售成本：{self.saleCostAmount}')    

    def updateCashSum(self):
        self.saleCashAmount = 0
        if self.incomeAmount.text().strip():
            self.saleCashAmount+= int(self.incomeAmount.text().strip())
        if self.taxAmount.text().strip():
            self.saleCashAmount += int(self.taxAmount.text().strip())
        # 在標籤中顯示總和
        self.labelCash.setText(f'現金：{self.saleCashAmount}')              

    def buttonClicked(self):

        # 檢查發票格式
        if not my_module.check_invoice_number_format(self.invoice.text().strip()) :
            QMessageBox.information(self, 'error', '發票號碼格式有誤!') 
            return  

        msg = ""
        msg += f'發票號碼: {self.invoice.text()} \n'
        # 打開 Excel 文件
        try:
            wb = openpyxl.load_workbook('112財報資料.xlsx', read_only=False)
        except FileNotFoundError:
            print('Excel 文件不存在')
            return

        # 選擇工作表 銷貨成本
        sheet = wb['銷貨成本']

        # 找到 A 欄中最後一個有值的單元格
        max_row = my_module.find_last_empty_row(sheet)

        # 寫入銷貨成本 分類帳
        max_row += 1
        sheet.cell(row=max_row, column=1).value = "-"
        sheet.cell(row=max_row, column=2).value = "銷貨成本"
        sheet.cell(row=max_row, column=4).value = self.saleCostAmount
        sheet.cell(row=max_row, column=7).value = self.invoice.text()
        msg += f'銷貨成本: {self.saleCostAmount}\n'


        # 選擇工作表 存貨
        sheet = wb['存貨']
        # 找到 A 欄中最後一個有值的單元格
        max_row = my_module.find_last_empty_row(sheet)
        # 寫入存貨 分類帳
        for i in range(5):
            if self.amount_edit[i].text().strip() and int(self.quantity[i].currentText()) > 0:
                max_row += 1
                sheet.cell(row=max_row, column=1).value = "-"
                sheet.cell(row=max_row, column=3).value = self.optionItems[i].currentText()
                sheet.cell(row=max_row, column=5).value = int(self.amount_edit[i].text())
                sheet.cell(row=max_row, column=7).value = self.invoice.text()
                msg += f'------------ {self.optionItems[i].currentText()},金額: {self.amount_edit[i].text()},數量: {self.quantity[i].currentText()}\n'
 

        # 選擇工作表 現金
        sheet = wb['現金']
        # 找到 A 欄中最後一個有值的單元格
        max_row = my_module.find_last_empty_row(sheet)
        # 寫入現金 分類帳 
        max_row += 1
        sheet.cell(row=max_row, column=1).value = "-"
        sheet.cell(row=max_row, column=3).value = "現金"
        sheet.cell(row=max_row, column=5).value = self.saleCashAmount
        sheet.cell(row=max_row, column=7).value = self.invoice.text()
        msg += f'現金: {self.saleCashAmount} \n'

        # 選擇工作表 銷貨收入
        sheet = wb['銷貨收入']
        # 找到 A 欄中最後一個有值的單元格
        max_row = my_module.find_last_empty_row(sheet)
        # 寫入銷貨收入 分類帳 
        if self.incomeAmount.text().strip() and int(self.incomeAmount.text()) > 0:
            max_row += 1
            sheet.cell(row=max_row, column=1).value = "-"
            sheet.cell(row=max_row, column=3).value = "銷貨收入"
            sheet.cell(row=max_row, column=5).value = int(self.incomeAmount.text())
            sheet.cell(row=max_row, column=7).value = self.invoice.text()
            msg += f'------------ 銷貨收入: {self.incomeAmount.text()}\n'

        # 選擇工作表 銷項稅額
        sheet = wb['銷項稅額']
        # 找到 A 欄中最後一個有值的單元格
        max_row = my_module.find_last_empty_row(sheet)
        # 寫入銷項稅額 分類帳
        if self.taxAmount.text().strip() and int(self.taxAmount.text()) > 0:
            max_row += 1
            sheet.cell(row=max_row, column=1).value = "-"
            sheet.cell(row=max_row, column=3).value = "銷項稅額"
            sheet.cell(row=max_row, column=5).value = int(self.taxAmount.text())
            sheet.cell(row=max_row, column=7).value = self.invoice.text()
            msg += f'------------ 銷項稅額: {self.taxAmount.text()}\n'


        
        msg += "請問是否確認存檔?"
        msg_box = QMessageBox()
        msg_box.setWindowTitle('確認')
        msg_box.setText(msg)
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        choice = msg_box.exec_()

        # If user clicks "Yes", save Excel file
        if choice == QMessageBox.No:
             return
        else:
            pass


        # 保存文件
        wb.save('112財報資料.xlsx')
        print('已將資料寫入 Excel 文件')
        QMessageBox.information(self, 'Save Excel', '資料寫入成功')
        self.resetData() # 重置資料

    # 重置資料
    def resetData(self):
        for i in range(5):
            self.optionItems[i].setCurrentIndex(0)
            self.amount_edit[i].clear()
            self.quantity[i].setCurrentIndex(0)
            self.taxAmount.clear()           
            self.incomeAmount.clear()
            self.taxAmount.clear()
